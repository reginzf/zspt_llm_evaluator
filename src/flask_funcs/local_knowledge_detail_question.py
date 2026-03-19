# 问题集管理相关路由
from flask import Blueprint, request, jsonify, send_file

import logging
import pandas as pd
import os
import tempfile
import uuid
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from src.flask_funcs.common_utils import validate_required_fields, get_knowledge_base_binding_info, handle_file_upload, \
    generate_unique_id

from src.sql_funs.questions_crud import QuestionsCRUD  # 导入新的QuestionsCRUD类

# 创建logger
logger = logging.getLogger(__name__)

# 创建蓝图
local_knowledge_question_bp = Blueprint('local_knowledge_question', __name__)

# 临时存储上传的导入数据 (内存存储，生产环境可考虑使用Redis)
_import_temp_storage = {}

# 问题类型有效值
VALID_QUESTION_TYPES = ['factual', 'contextual', 'conceptual', 'reasoning', 'application']


@local_knowledge_question_bp.route('/local_knowledge_detail/question_set/create', methods=['POST'])
@local_knowledge_question_bp.route('/api/local_knowledge_detail/question_set/create', methods=['POST'])
def create_question_set():
    """创建问题集"""
    try:
        data = request.get_json()
        required_fields = ['knowledge_id', 'set_name', 'set_type']
        missing_field = validate_required_fields(data, required_fields)
        if missing_field:
            return jsonify({'success': False, 'message': f'缺少必要字段: {missing_field}'}), 400

        knowledge_id = data['knowledge_id']
        set_name = data['set_name']
        set_type = data['set_type']

        question_set_id = generate_unique_id('qs', 8)

        # 创建问题集
        with QuestionsCRUD() as crud:
            result = crud.question_config_create(
                question_id=question_set_id,
                question_name=set_name,
                knowledge_id=knowledge_id,
                question_set_type=set_type
            )

        if result:
            return jsonify({
                'success': True,
                'message': '问题集创建成功',
                'data': {'question_set_id': question_set_id}
            })
        else:
            return jsonify({
                'success': False,
                'message': '问题集创建失败'
            }), 500

    except Exception as e:
        logger.error(f"创建问题集时发生错误: {str(e)}")
        return jsonify({'success': False, 'message': f'创建问题集时发生错误: {str(e)}'}), 500


@local_knowledge_question_bp.route('/local_knowledge_detail/question/set/list', methods=['GET'])
@local_knowledge_question_bp.route('/api/local_knowledge_detail/question/set/list', methods=['GET'])
def get_question_set_list():
    """获取问题集列表"""
    try:
        knowledge_id = request.args.get('knowledge_id')

        if not knowledge_id:
            return jsonify({'success': False, 'message': '缺少knowledge_id参数'}), 400

        with QuestionsCRUD() as crud:
            question_sets = crud.question_config_list(knowledge_id=knowledge_id)

        if question_sets:
            # 转换数据格式
            result = []
            for qs in question_sets:
                result.append(crud._question_config_to_json(qs))
            return jsonify({
                'success': True,
                'data': result
            })
        else:
            return jsonify({
                'success': True,
                'data': []
            })

    except Exception as e:
        logger.error(f"获取问题集列表时发生错误: {str(e)}")
        return jsonify({'success': False, 'message': f'获取问题集列表时发生错误: {str(e)}'}), 500


@local_knowledge_question_bp.route('/local_knowledge_detail/question/set/detail', methods=['GET'])
@local_knowledge_question_bp.route('/api/local_knowledge_detail/question/set/detail', methods=['GET'])
def get_question_set_detail():
    """获取问题集详情"""
    try:
        set_id = request.args.get('set_id')

        if not set_id:
            return jsonify({'success': False, 'message': '缺少set_id参数'}), 400

        with QuestionsCRUD() as crud:
            question_sets = crud.question_config_list(question_id=set_id)

        if question_sets and len(question_sets) > 0:
            result = crud._question_config_to_json(question_sets[0])
            return jsonify({
                'success': True,
                'data': result
            })
        else:
            return jsonify({
                'success': False,
                'message': '未找到问题集'
            }), 404

    except Exception as e:
        logger.error(f"获取问题集详情时发生错误: {str(e)}")
        return jsonify({'success': False, 'message': f'获取问题集详情时发生错误: {str(e)}'}), 500


@local_knowledge_question_bp.route('/local_knowledge_detail/question/set/delete', methods=['DELETE'])
@local_knowledge_question_bp.route('/api/local_knowledge_detail/question/set/delete', methods=['DELETE'])
def delete_question_set():
    """删除问题集"""
    try:
        # 支持从 query 参数或 body 中获取 set_id/question_id
        set_id = request.args.get('set_id') or request.args.get('question_id')
        if not set_id:
            data = request.get_json(silent=True) or {}
            set_id = data.get('set_id') or data.get('question_id')

        if not set_id:
            return jsonify({'success': False, 'message': '缺少必要字段: set_id 或 question_id'}), 400

        # 删除问题集（注意：这里应该考虑级联删除相关的问题）
        with QuestionsCRUD() as crud:
            # 先删除相关的问题（从各个问题表中删除）
            # 为了简单起见，我们先只删除问题集配置记录
            # 实际应用中可能需要更复杂的级联删除逻辑
            result = crud.question_config_delete(set_id)

        if result:
            return jsonify({
                'success': True,
                'message': '问题集删除成功'
            })
        else:
            return jsonify({
                'success': False,
                'message': '问题集删除失败或问题集不存在'
            }), 500

    except Exception as e:
        logger.error(f"删除问题集时发生错误: {str(e)}")
        return jsonify({'success': False, 'message': f'删除问题集时发生错误: {str(e)}'}), 500


@local_knowledge_question_bp.route('/local_knowledge_detail/question/set/update', methods=['PUT'])
@local_knowledge_question_bp.route('/api/local_knowledge_detail/question/set/update', methods=['PUT'])
def update_question_set():
    """更新问题集信息"""
    try:
        data = request.get_json()
        required_fields = ['set_id']
        missing_field = validate_required_fields(data, required_fields)
        if missing_field:
            return jsonify({'success': False, 'message': f'缺少必要字段: {missing_field}'}), 400

        set_id = data['set_id']
        set_name = data.get('set_name')
        set_type = data.get('set_type')  # 对应数据库中的question_set_type

        # 更新问题集
        with QuestionsCRUD() as crud:
            result = crud.question_config_update(
                question_id=set_id,
                question_name=set_name,
                question_set_type=set_type  # 添加对question_set_type的更新
            )

        if result:
            return jsonify({
                'success': True,
                'message': '问题集更新成功'
            })
        else:
            return jsonify({
                'success': False,
                'message': '问题集更新失败或问题集不存在'
            }), 500

    except Exception as e:
        logger.error(f"更新问题集时发生错误: {str(e)}")
        return jsonify({'success': False, 'message': f'更新问题集时发生错误: {str(e)}'}), 500


@local_knowledge_question_bp.route('/local_knowledge_detail/question/create', methods=['POST'])
@local_knowledge_question_bp.route('/api/local_knowledge_detail/question/create', methods=['POST'])
def create_question():
    """创建问题"""
    try:
        data = request.get_json()
        required_fields = ['set_id', 'question_set_type', 'question_type', 'question_content']
        missing_field = validate_required_fields(data, required_fields)
        if missing_field:
            return jsonify({'success': False, 'message': f'缺少必要字段: {missing_field}'}), 400

        set_id = data['set_id']
        question_set_type = data['question_set_type']
        question_type = data['question_type']
        question_content = data['question_content']
        chunk_ids = data.get('chunk_ids', [])

        # 处理 chunk_ids：如果是字符串（如 '123,456'），按逗号分割为列表
        if isinstance(chunk_ids, str):
            chunk_ids = [cid.strip() for cid in chunk_ids.split(',') if cid.strip()]
        elif not isinstance(chunk_ids, list):
            chunk_ids = []

        # 生成问题ID
        question_id = generate_unique_id('qst', 8)

        question_data = {"question_id": question_id,
                "question_set_id": set_id,
                "question_type": question_type,
                "question_content": question_content,
                "chunk_ids": chunk_ids}
        # 根据问题类型存储到相应表
        result = None
        with QuestionsCRUD() as crud:
            # 将问题写入到对应表中
            if question_set_type == 'basic':
                result = crud.create_basic_question(**question_data)
            elif question_set_type == 'detailed':
                result = crud.create_detailed_question(**question_data)
            elif question_set_type == 'mechanism':
                result = crud.create_mechanism_question(**question_data)
            elif question_set_type == 'thematic':
                result = crud.create_thematic_question(**question_data)
            
            if result:
                # 增加问题集的问题计数（原子操作）
                import psycopg2
                try:
                    # 直接在数据库层面增加计数，避免竞态条件
                    update_query = "UPDATE ai_question_config SET question_count = question_count + 1, updated_at = CURRENT_TIMESTAMP WHERE question_id = %s"
                    crud.cursor.execute(update_query, (set_id,))
                    crud.connection.commit()
                except psycopg2.Error as db_error:
                    crud.connection.rollback()
                    logger.error(f"更新问题计数失败: {db_error}")
                    
        if result:
            return jsonify({
                'success': True,
                'message': '问题创建成功',
                'data': {'question_id': question_id}
            })
        else:
            return jsonify({
                'success': False,
                'message': '问题创建失败'
            }), 500

    except Exception as e:
        logger.error(f"创建问题时发生错误: {str(e)}")
        return jsonify({'success': False, 'message': f'创建问题时发生错误: {str(e)}'}), 500


@local_knowledge_question_bp.route('/local_knowledge_detail/question/list', methods=['POST'])
@local_knowledge_question_bp.route('/api/local_knowledge_detail/question/list', methods=['POST'])
def get_question_list():
    """获取问题列表 (新版本，使用JSON格式)"""
    try:
        data = request.get_json()
        set_id = data.get('set_id')
        question_type = data.get('question_type', 'basic')  # 默认为basic

        if not set_id:
            return jsonify({'success': False, 'message': '缺少set_id参数'}), 400

        with QuestionsCRUD() as crud:
            # 根据问题类型获取问题列表 - 修改调用方式
            questions = crud.get_questions_by_type(
                question_set_type=question_type,
                question_set_id=set_id
            )

        if questions:
            # 转换数据格式 - 根据实际表结构调整
            result = []
            for q in questions:
                # 根据实际表结构调整字段索引
                # ai_basic_questions, ai_detailed_questions, ai_mechanism_questions, ai_thematic_questions 表结构:
                # id, question_id, question_set_id, question_type, question_content, chunk_ids, created_at, updated_at
                question_obj = {
                    'id': q[0] if len(q) > 0 else None,
                    'question_id': q[1] if len(q) > 1 else None,
                    'question_set_id': q[2] if len(q) > 2 else None,
                    'question_type': q[3] if len(q) > 3 else None,
                    'question_content': q[4] if len(q) > 4 else None,
                    'chunk_ids': q[5] if len(q) > 5 else [],
                    'created_at': q[6].isoformat() if len(q) > 6 and q[6] else None,
                    'updated_at': q[7].isoformat() if len(q) > 7 and q[7] else None
                }
                result.append(question_obj)
            return jsonify({
                'success': True,
                'data': result
            })
        else:
            return jsonify({
                'success': True,
                'data': []
            })

    except Exception as e:
        logger.error(f"获取问题列表时发生错误: {str(e)}")
        return jsonify({'success': False, 'message': f'获取问题列表时发生错误: {str(e)}'}), 500


@local_knowledge_question_bp.route('/local_knowledge_detail/question/detail', methods=['POST'])
@local_knowledge_question_bp.route('/api/local_knowledge_detail/question/detail', methods=['POST'])
def get_question_detail():
    """获取问题详情 (新版本，使用JSON格式)"""
    try:
        data = request.get_json()
        question_id = data.get('question_id')
        question_set_type = data.get('question_set_type')

        if not question_id or not question_set_type:
            return jsonify({'success': False, 'message': '缺少question_id或question_set_type参数'}), 400

        with QuestionsCRUD() as crud:
            # 获取指定类型的问题 - 修改调用方式
            questions = crud.get_questions_by_type(
                question_set_type=question_set_type,
                question_id=question_id
            )

        if questions and len(questions) > 0:
            q = questions[0]
            # 根据实际表结构调整字段索引
            result = {
                'id': q[0] if len(q) > 0 else None,
                'question_id': q[1] if len(q) > 1 else None,
                'question_set_id': q[2] if len(q) > 2 else None,
                'question_type': q[3] if len(q) > 3 else None,
                'question_content': q[4] if len(q) > 4 else None,
                'chunk_ids': q[5] if len(q) > 5 else [],
                'created_at': q[6].isoformat() if len(q) > 6 and q[6] else None,
                'updated_at': q[7].isoformat() if len(q) > 7 and q[7] else None
            }
            return jsonify({
                'success': True,
                'data': result
            })
        else:
            return jsonify({
                'success': False,
                'message': '未找到问题'
            }), 404

    except Exception as e:
        logger.error(f"获取问题详情时发生错误: {str(e)}")
        return jsonify({'success': False, 'message': f'获取问题详情时发生错误: {str(e)}'}), 500


@local_knowledge_question_bp.route('/local_knowledge_detail/question/update', methods=['PUT'])
@local_knowledge_question_bp.route('/api/local_knowledge_detail/question/update', methods=['PUT'])
def update_question():
    """更新问题"""
    try:
        data = request.get_json()
        required_fields = ['question_id', 'question_type']
        missing_field = validate_required_fields(data, required_fields)
        if missing_field:
            return jsonify({'success': False, 'message': f'缺少必要字段: {missing_field}'}), 400

        question_id = data['question_id']
        question_type = data['question_type']
        question_content = data.get('question_content')
        chunk_ids = data.get('chunk_ids')

        # 处理 chunk_ids：如果是字符串（如 '123,456'），按逗号分割为列表
        if isinstance(chunk_ids, str):
            chunk_ids = [cid.strip() for cid in chunk_ids.split(',') if cid.strip()]
        elif chunk_ids is not None and not isinstance(chunk_ids, list):
            chunk_ids = []

        # 使用CRUD类的更新方法
        with QuestionsCRUD() as crud:
            result = crud.update_question_by_type(
                question_id=question_id,
                question_type=question_type,
                question_content=question_content,
                chunk_ids=chunk_ids
            )

            if result:
                return jsonify({
                    'success': True,
                    'message': '问题更新成功'
                })
            else:
                return jsonify({
                    'success': False,
                    'message': '问题不存在或未被更新'
                }), 404

    except Exception as e:
        logger.error(f"更新问题时发生错误: {str(e)}")
        return jsonify({'success': False, 'message': f'更新问题时发生错误: {str(e)}'}), 500


@local_knowledge_question_bp.route('/local_knowledge_detail/question/delete', methods=['DELETE'])
@local_knowledge_question_bp.route('/api/local_knowledge_detail/question/delete', methods=['DELETE'])
def delete_question():
    """删除问题"""
    try:
        data = request.get_json()
        required_fields = ['question_id', 'question_set_type']
        missing_field = validate_required_fields(data, required_fields)
        if missing_field:
            return jsonify({'success': False, 'message': f'缺少必要字段: {missing_field}'}), 400

        question_id = data['question_id']
        question_set_type = data['question_set_type']

        # 根据问题类型从相应表中删除
        with QuestionsCRUD() as crud:
            result = crud.delete_question_by_type(question_id, question_set_type)

            if result:
                return jsonify({
                    'success': True,
                    'message': '问题删除成功'
                })
            else:
                return jsonify({
                    'success': False,
                    'message': '问题不存在或删除失败'
                }), 404

    except Exception as e:
        logger.error(f"删除问题时发生错误: {str(e)}")
        return jsonify({'success': False, 'message': f'删除问题时发生错误: {str(e)}'}), 500


@local_knowledge_question_bp.route('/api/local_knowledge_detail/question/import/upload', methods=['POST'])
def upload_question_import():
    """上传并预览问题导入文件"""
    try:
        # 检查文件
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '未找到上传的文件'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '未选择文件'}), 400

        # 检查文件格式
        if not file.filename.endswith('.xlsx'):
            return jsonify({'success': False, 'message': '请上传.xlsx格式的文件'}), 400

        # 检查文件大小 (10MB限制)
        file.seek(0, 2)  # 移动到文件末尾
        file_size = file.tell()
        file.seek(0)  # 重置到文件开头
        if file_size > 10 * 1024 * 1024:
            return jsonify({'success': False, 'message': '文件大小超过10MB限制'}), 413

        # 保存到临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            file.save(tmp_file.name)
            tmp_path = tmp_file.name

        try:
            # 使用pandas读取Excel
            df = pd.read_excel(tmp_path)

            # 检查Sheet是否为空
            if df.empty:
                return jsonify({'success': False, 'message': 'Excel文件为空'}), 400

            # 检查必填列
            required_columns = ['question_type', 'question_content']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return jsonify({
                    'success': False,
                    'message': f'缺少必填列: {", ".join(missing_columns)}'
                }), 400

            # 逐行验证和解析
            preview = []
            valid_count = 0
            invalid_count = 0

            for idx, row in df.iterrows():
                row_index = idx + 2  # Excel行号从2开始（考虑表头）
                row_data = {
                    'row_index': row_index,
                    'question_type': str(row.get('question_type', '')).strip(),
                    'question_content': str(row.get('question_content', '')).strip(),
                    'chunk_ids': [],
                    'is_valid': True,
                    'error_msg': None
                }

                # 验证question_type
                if not row_data['question_type']:
                    row_data['is_valid'] = False
                    row_data['error_msg'] = '问题类型不能为空'
                elif row_data['question_type'] not in VALID_QUESTION_TYPES:
                    row_data['is_valid'] = False
                    row_data['error_msg'] = f"问题类型无效，可选值: {', '.join(VALID_QUESTION_TYPES)}"

                # 验证question_content
                if not row_data['question_content']:
                    row_data['is_valid'] = False
                    if row_data['error_msg']:
                        row_data['error_msg'] += '; 问题内容不能为空'
                    else:
                        row_data['error_msg'] = '问题内容不能为空'

                # 解析chunk_ids
                chunk_ids_str = str(row.get('chunk_ids', '')).strip()
                if chunk_ids_str and chunk_ids_str != 'nan':
                    row_data['chunk_ids'] = [cid.strip() for cid in chunk_ids_str.split(',') if cid.strip()]

                if row_data['is_valid']:
                    valid_count += 1
                else:
                    invalid_count += 1

                preview.append(row_data)

            # 生成import_token
            import_token = f"import_{uuid.uuid4().hex[:16]}"

            # 存储临时数据（30分钟过期）
            _import_temp_storage[import_token] = {
                'preview': preview,
                'temp_file': tmp_path,
                'created_at': datetime.now(),
                'expires_at': datetime.now() + timedelta(minutes=30)
            }

            return jsonify({
                'success': True,
                'data': {
                    'import_token': import_token,
                    'preview': preview[:5],  # 只返回前5行用于预览
                    'total_count': len(preview),
                    'valid_count': valid_count,
                    'invalid_count': invalid_count
                }
            })

        except pd.errors.EmptyDataError:
            return jsonify({'success': False, 'message': 'Excel文件为空'}), 400
        except Exception as e:
            logger.error(f"解析Excel文件时发生错误: {str(e)}")
            return jsonify({'success': False, 'message': f'解析Excel文件失败: {str(e)}'}), 500

    except Exception as e:
        logger.error(f"上传导入文件时发生错误: {str(e)}")
        return jsonify({'success': False, 'message': f'上传文件时发生错误: {str(e)}'}), 500


@local_knowledge_question_bp.route('/api/local_knowledge_detail/question/import/confirm', methods=['POST'])
def confirm_question_import():
    """确认导入问题"""
    try:
        data = request.get_json()
        required_fields = ['import_token', 'set_id', 'set_type']
        missing_field = validate_required_fields(data, required_fields)
        if missing_field:
            return jsonify({'success': False, 'message': f'缺少必要字段: {missing_field}'}), 400

        import_token = data['import_token']
        set_id = data['set_id']
        set_type = data['set_type']

        # 验证set_type
        valid_set_types = ['basic', 'detailed', 'mechanism', 'thematic']
        if set_type not in valid_set_types:
            return jsonify({
                'success': False,
                'message': f'无效的问题集类型，可选值: {", ".join(valid_set_types)}'
            }), 400

        # 获取临时数据
        if import_token not in _import_temp_storage:
            return jsonify({'success': False, 'message': '导入令牌已过期或无效，请重新上传文件'}), 400

        temp_data = _import_temp_storage[import_token]

        # 检查是否过期
        if datetime.now() > temp_data['expires_at']:
            # 清理过期数据
            _cleanup_import_temp(import_token)
            return jsonify({'success': False, 'message': '导入令牌已过期，请重新上传文件'}), 400

        preview = temp_data['preview']

        # 过滤掉无效行
        valid_rows = [row for row in preview if row['is_valid']]

        if not valid_rows:
            return jsonify({'success': False, 'message': '没有有效的问题数据可以导入'}), 400

        # 批量插入问题
        inserted_count = 0
        failed_count = 0

        with QuestionsCRUD() as crud:
            # 确定创建问题的函数
            create_func = None
            if set_type == 'basic':
                create_func = crud.create_basic_question
            elif set_type == 'detailed':
                create_func = crud.create_detailed_question
            elif set_type == 'mechanism':
                create_func = crud.create_mechanism_question
            elif set_type == 'thematic':
                create_func = crud.create_thematic_question

            for row in valid_rows:
                try:
                    question_id = generate_unique_id('qst', 8)
                    result = create_func(
                        question_id=question_id,
                        question_type=row['question_type'],
                        question_content=row['question_content'],
                        question_set_id=set_id,
                        chunk_ids=row['chunk_ids']
                    )

                    if result:
                        inserted_count += 1
                        # 更新问题集计数
                        try:
                            update_query = "UPDATE ai_question_config SET question_count = question_count + 1, updated_at = CURRENT_TIMESTAMP WHERE question_id = %s"
                            crud.cursor.execute(update_query, (set_id,))
                            crud.connection.commit()
                        except Exception as e:
                            crud.connection.rollback()
                            logger.error(f"更新问题计数失败: {e}")
                    else:
                        failed_count += 1
                        logger.error(f"创建问题失败: {row}")
                except Exception as e:
                    failed_count += 1
                    logger.error(f"插入问题时发生错误: {str(e)}, row: {row}")

        # 清理临时数据
        _cleanup_import_temp(import_token)

        return jsonify({
            'success': True,
            'data': {
                'inserted_count': inserted_count,
                'failed_count': failed_count
            }
        })

    except Exception as e:
        logger.error(f"确认导入时发生错误: {str(e)}")
        return jsonify({'success': False, 'message': f'导入时发生错误: {str(e)}'}), 500


def _cleanup_import_temp(import_token: str):
    """清理临时导入数据"""
    if import_token in _import_temp_storage:
        temp_data = _import_temp_storage[import_token]
        # 删除临时文件
        try:
            if 'temp_file' in temp_data and os.path.exists(temp_data['temp_file']):
                os.remove(temp_data['temp_file'])
        except Exception as e:
            logger.error(f"删除临时文件失败: {e}")
        # 删除内存中的数据
        del _import_temp_storage[import_token]


@local_knowledge_question_bp.route('/api/local_knowledge_detail/question/import/template', methods=['GET'])
def download_question_template():
    """下载问题导入模板"""
    try:
        # 创建Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "问题导入模板"

        # 设置表头
        headers = ['question_type', 'question_content', 'chunk_ids']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            # 设置表头样式（加粗）
            cell.font = cell.font.copy(bold=True)

        # 添加示例数据
        sample_data = [
            ['factual', '什么是OSPF协议？', 'chunk1,chunk2'],
            ['contextual', '根据文档内容，配置步骤是什么？', ''],
            ['conceptual', '如何理解路由表的工作原理？', 'chunk3']
        ]
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 为 question_type 列创建下拉列表（第2行到第1000行）
        dv = DataValidation(
            type="list",
            formula1='"factual,contextual,conceptual,reasoning,application"',
            allow_blank=True
        )
        dv.error = "请选择有效的问题类型"
        dv.errorTitle = "无效输入"
        dv.prompt = "请选择问题类型"
        dv.promptTitle = "问题类型"
        ws.add_data_validation(dv)
        dv.add('A2:A1000')  # question_type 列，从第2行到第1000行

        # 设置列宽
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25

        # 创建临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_path = tmp_file.name
            wb.save(tmp_path)

        # 发送文件
        response = send_file(
            tmp_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            attachment_filename='question_import_template.xlsx'
        )

        # 添加回调删除临时文件
        @response.call_on_close
        def cleanup():
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception as e:
                logger.error(f"删除模板临时文件失败: {e}")

        return response

    except Exception as e:
        logger.error(f"生成模板文件时发生错误: {str(e)}")
        return jsonify({'success': False, 'message': f'生成模板失败: {str(e)}'}), 500
